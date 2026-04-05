import subprocess
import os
import sys
import time
import json
import re
from datetime import datetime
from typing import Optional
from pathlib import Path

# Ensure src directory is in Python path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import PROJECT_ROOT, OPENCODE_CMD


def call_opencode(prompt: str, continue_session_id: Optional[str] = None) -> tuple:
    """调用 OpenCode 并返回 (output, session_id)

    Args:
        prompt: 提示词
        continue_session_id: 可选的 session ID，用于继续对话

    Returns:
        tuple: (output文本, session_id)
    """
    prompt_file = os.path.join(PROJECT_ROOT, ".prompt.txt")
    with open(prompt_file, "w", encoding="utf-8") as f:
        f.write(prompt)

    try:
        cmd_parts = [OPENCODE_CMD, "run", f"@{prompt_file}", "--format", "json"]

        # 如果指定了要继续的 session，使用 --session 参数
        if continue_session_id:
            cmd_parts.extend(["--session", continue_session_id])

        cmd = subprocess.list2cmdline(cmd_parts)
        try:
            result = subprocess.run(
                cmd,
                cwd=PROJECT_ROOT,
                capture_output=True,
                timeout=300,  # 增加到 5 分钟
                shell=True
            )
        except subprocess.TimeoutExpired:
            return ("OpenCode execution timed out after 5 minutes", None)

        # 解析 JSON 输出（包含 session ID）
        session_id = None
        output_text = ""

        if result.stdout:
            try:
                # JSON 格式输出时，每行是一个 JSON 事件
                for line in result.stdout.decode('utf-8', errors='replace').strip().split('\n'):
                    if line.strip():
                        try:
                            event = json.loads(line)
                            # 从第一个事件获取 session ID
                            if event.get('sessionID') and not session_id:
                                session_id = event['sessionID']
                            # 从 text 事件获取输出文本
                            if event.get('type') == 'text' and 'text' in event.get('part', {}):
                                output_text += event['part']['text']
                        except json.JSONDecodeError:
                            pass
            except Exception as e:
                # JSON 解析失败，回退到原始输出
                pass

        # 如果没有从 JSON 获取到 session ID，尝试从纯文本输出中提取
        if not session_id and result.stdout:
            stdout_text = result.stdout.decode('utf-8', errors='replace')
            session_id_match = re.search(r'sessionID["\s:]+([a-zA-Z0-9_]+)', stdout_text)
            if session_id_match:
                session_id = session_id_match.group(1)

        # 如果仍然没有输出文本，使用原始输出
        if not output_text and result.stdout:
            output_text = result.stdout.decode('utf-8', errors='replace')[:3000]

        return output_text, session_id

    finally:
        time.sleep(1)
        if os.path.exists(prompt_file):
            try:
                os.remove(prompt_file)
            except:
                pass


def generate_excel(
    url: str,
    filepath: str,
    description: str,
    username: str = "",
    password: str = "",
    continue_excel: Optional[str] = None,
    continue_session_id: Optional[str] = None,
) -> dict:
    """生成 Excel 测试用例（非交互式）

    Args:
        url: UI 地址
        filepath: 保存路径
        description: 测试描述
        username: 登录账号（可选）
        password: 登录密码（可选）
        continue_excel: 可选的 Excel 文件路径，用于继续之前的 session
        continue_session_id: 可选的 session ID（由 API 层从 TaskDB 查询后传入）

    Returns:
        dict: {"status": "success"/"warning"/"error", "message": str, "output": str}
    """
    if not url:
        return {"status": "error", "message": "URL 不能为空", "output": ""}
    if not filepath:
        return {"status": "error", "message": "保存路径不能为空", "output": ""}
    if not description:
        return {"status": "error", "message": "测试描述不能为空", "output": ""}

    login_info = f"登录账号：{username}，密码：{password}" if username else "无需登录"

    prompt = f"""请为以下网站生成 Excel 测试用例：
网站：{url}
{login_info}
测试需求：{description}
保存路径：{filepath}

Excel 格式要求：
- 用例编号（TC-序号）
- 用例名称
- 模块/功能
- 优先级（P0/P1/P2）
- 预置条件
- 测试步骤
- 预期结果

使用 openpyxl 生成 Excel，保存到 {filepath}

重要：生成完成后，必须使用 openpyxl 将以下配置信息写入 Excel 的"配置"工作表：
1. URL：{url}
2. 登录账号：{username or '无'}
3. 登录密码：{password or '无'}
4. 测试描述：{description}

如果生成了 Python 脚本文件，必须保存到 templates/ 目录下，不要直接在当前目录生成脚本文件。"""

    print(f"[INFO] 正在执行 OpenCode...")

    output, session_id = call_opencode(prompt, continue_session_id)

    print(f"[INFO] Session ID from OpenCode: {session_id}")

    excel_path = os.path.join(PROJECT_ROOT, filepath)

    # Handle directory path - find .xlsx file inside
    actual_file_path = excel_path
    if os.path.isdir(excel_path):
        for f in os.listdir(excel_path):
            if f.endswith('.xlsx'):
                actual_file_path = os.path.join(excel_path, f)
                break
        else:
            return {"status": "warning", "message": f"目录中未找到 Excel 文件: {filepath}", "output": output, "session_id": session_id}

    if os.path.exists(actual_file_path):
        return {"status": "success", "message": f"Excel 已生成: {filepath}", "output": output, "actual_file": actual_file_path, "session_id": session_id}
    else:
        return {"status": "warning", "message": f"Excel 可能未生成，请检查: {filepath}", "output": output}


def continue_session(excel_file: str, save_path: str = "", session_id: str = None) -> dict:
    """继续 session，读取 Excel 生成测试代码（非交互式）

    Args:
        excel_file: Excel 文件路径
        save_path: 保存路径（可选）
        session_id: 可选的 session ID（由 API 层从 TaskDB 查询后传入）

    Returns:
        dict: {"status": "success"/"error", "message": str, "output": str}
    """
    if not excel_file:
        return {"status": "error", "message": "Excel 文件路径不能为空", "output": ""}

    excel_path = os.path.join(PROJECT_ROOT, excel_file)

    # Handle directory path - find .xlsx file inside
    if os.path.isdir(excel_path):
        for f in os.listdir(excel_path):
            if f.endswith('.xlsx'):
                excel_path = os.path.join(excel_path, f)
                break
        else:
            return {"status": "error", "message": f"目录中未找到 Excel 文件: {excel_file}", "output": ""}

    if not os.path.exists(excel_path):
        return {"status": "error", "message": f"Excel 文件不存在: {excel_file}", "output": ""}

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)

    config = {}
    if "配置" in wb.sheetnames:
        ws_config = wb["配置"]
        for row in ws_config.iter_rows(min_row=2, values_only=True):
            if row[0]:
                config[row[0]] = row[1] if len(row) > 1 else ""

    if "Sheet" in wb.sheetnames:
        ws = wb["Sheet"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    test_cases = rows[1:] if len(rows) > 1 else []

    test_summary = "\n".join([f"- {tc[0]}: {tc[1]}" for tc in test_cases[:10]])

    url = config.get("URL", "")
    username = config.get("登录账号", "") if config.get("登录账号", "") != "无" else ""
    password = config.get("登录密码", "") if config.get("登录密码", "") != "无" else ""

    # Use save_path as folder name, fallback to parsing from excel_file
    folder_name = save_path
    if not folder_name:
        folder_name = os.path.basename(excel_file.rstrip('/'))
        if not folder_name:
            folder_name = os.path.basename(os.path.dirname(excel_file.rstrip('/')))
    else:
        # save_path might be "test_cases/google/" or similar - extract the last component
        folder_name = folder_name.rstrip('/')
        if '/' in folder_name or '\\' in folder_name:
            folder_name = os.path.basename(folder_name)

    # Extract the last component of folder_name for the file name
    file_name = os.path.basename(folder_name.rstrip('/'))

    prompt = f"""请读取 Excel 文件 {excel_file} 中的测试用例，为其生成 pytest 测试代码。

配置信息：
- URL：{url}
- 登录账号：{username or '无'}
- 登录密码：{password or '无'}

测试用例摘要：
{test_summary}

请生成：
1. Page Object 类（封装页面元素和操作），使用 Playwright
2. pytest 测试用例，使用配置表中的 URL 和账号密码

重要：目录 tests/{folder_name}/ 已经创建好了，必须将测试代码写入该目录：
- Page Object 文件：tests/{folder_name}/pages.py
- 测试用例文件：tests/{folder_name}/test_{file_name}.py

不要写到其他位置，只写到这个已存在的目录中。"""

    # Create the target directory if it doesn't exist
    target_dir = os.path.join(PROJECT_ROOT, "tests", folder_name)
    os.makedirs(target_dir, exist_ok=True)
    print(f"[INFO] 创建目录: {target_dir}")

    print(f"[INFO] 正在执行 OpenCode...")

    output, new_session_id = call_opencode(prompt, session_id)

    print(f"[INFO] New session ID from OpenCode: {new_session_id}")

    test_code_dir = "tests/" + folder_name + "/test_" + file_name.replace(" ", "_") + ".py"

    return {
        "status": "success",
        "message": "测试代码生成完成",
        "output": output,
        "test_code_dir": test_code_dir,
        "session_id": new_session_id,
    }



# CLI 入口函数
def main():
    """CLI 入口，交互式生成"""
    import argparse
    parser = argparse.ArgumentParser(description='OpenCode Session Manager')
    parser.add_argument('--generate', action='store_true', help='生成 Excel 测试用例')
    parser.add_argument('--continue', dest='continue_file', metavar='EXCEL_FILE',
                        help='继续 session，读取 Excel 生成测试代码')
    args = parser.parse_args()

    if args.generate:
        url = input("UI 地址: ").strip()
        filepath = input("保存路径: ").strip()
        username = input("登录账号（可选）: ").strip()
        password = input("登录密码（可选）: ").strip()
        description = input("测试描述: ").strip()

        result = generate_excel(url, filepath, description, username, password)
        print(result["message"])

    elif args.continue_file:
        result = continue_session(args.continue_file)
        print(result["message"])

    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
