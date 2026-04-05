import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "测试用例"

headers = ["用例编号", "用例名称", "模块/功能", "优先级", "预置条件", "测试步骤", "预期结果"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

test_cases = [
    ["TC-001", "页面加载验证", "页面功能", "P0", "网络连接正常", "1. 打开浏览器\n2. 访问 https://example.com\n3. 等待页面加载完成", "页面正常加载，显示内容"],
    ["TC-002", "页面标题验证", "页面功能", "P0", "已访问页面", "1. 访问 https://example.com\n2. 获取页面标题", "页面标题正确显示"],
    ["TC-003", "页面内容验证", "页面功能", "P0", "已打开页面", "1. 访问 https://example.com\n2. 检查页面主体内容", "页面主体内容正常显示"],
    ["TC-004", "页面图片验证", "页面功能", "P1", "已打开页面", "1. 访问 https://example.com\n2. 检查页面中的图片", "图片正常加载显示"],
    ["TC-005", "页面链接验证", "页面功能", "P1", "已打开页面", "1. 访问 https://example.com\n2. 检查页面中的链接", "链接正常显示且可点击"],
    ["TC-006", "页面响应时间", "页面功能", "P2", "网络连接正常", "1. 记录开始时间\n2. 访问 https://example.com\n3. 记录页面加载完成时间", "页面加载时间在合理范围内"],
    ["TC-007", "页面兼容性", "页面功能", "P2", "网络连接正常", "1. 使用不同浏览器访问\n2. 检查页面显示", "各浏览器页面显示正常"],
]

for row_idx, case in enumerate(test_cases, 2):
    for col_idx, value in enumerate(case, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 35
ws.column_dimensions["G"].width = 30

ws_config = wb.create_sheet(title="配置")

config_data = [
    ("URL", "https://example.com"),
    ("登录账号", "无"),
    ("登录密码", "无"),
    ("测试描述", "test"),
]

for row_idx, (key, value) in enumerate(config_data, 1):
    ws_config.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
    ws_config.cell(row=row_idx, column=2, value=value)

ws_config.column_dimensions["A"].width = 15
ws_config.column_dimensions["B"].width = 30

wb.save("test3.xlsx")
print("Excel 文件已生成: test3.xlsx")
