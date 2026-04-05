from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "测试用例"

ws.append(["用例编号", "用例名称", "模块/功能", "优先级", "预置条件", "测试步骤", "预期结果"])

test_cases = [
    ["TC-1", "验证网站可访问", "网站访问", "P0", "网络正常", "1. 打开浏览器\n2. 访问 https://test.com", "页面正常加载"],
    ["TC-2", "验证页面标题", "网站访问", "P1", "网站可访问", "1. 访问 https://test.com\n2. 获取页面标题", "标题显示正确"],
    ["TC-3", "验证页面元素", "网站访问", "P1", "网站可访问", "1. 访问 https://test.com\n2. 检查页面关键元素", "关键元素正常显示"],
]

for tc in test_cases:
    ws.append(tc)

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column].width = max_length + 2

ws_config = wb.create_sheet("配置")
ws_config.append(["配置项", "值"])
ws_config.append(["URL", "https://test.com"])
ws_config.append(["登录账号", "无"])
ws_config.append(["登录密码", "无"])
ws_config.append(["测试描述", "test"])

for col in ws_config.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws_config.column_dimensions[column].width = max_length + 2

wb.save("test.xlsx")
print("Excel 文件已生成: test.xlsx")
