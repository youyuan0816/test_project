import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def create_google_home_test_cases():
    wb = openpyxl.Workbook()

    ws_cases = wb.active
    ws_cases.title = "测试用例"

    headers = ["用例编号", "用例名称", "模块/功能", "优先级", "预置条件", "测试步骤", "预期结果"]
    for col, header in enumerate(headers, 1):
        cell = ws_cases.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    test_cases = [
        ["TC-001", "验证Google首页正常加载", "首页", "P0", "网络正常", "1. 打开浏览器\n2. 访问Google首页\n3. 等待页面加载完成", "页面正常加载，显示Google搜索框"],
        ["TC-002", "验证页面标题显示正确", "首页", "P0", "页面正常加载", "1. 访问Google首页\n2. 获取页面标题", "页面标题显示为Google"],
        ["TC-003", "验证Google Logo显示", "首页", "P0", "页面正常加载", "1. 访问Google首页\n2. 检查Google Logo是否显示", "Google Logo正常显示"],
        ["TC-004", "验证搜索输入框存在且可编辑", "搜索", "P0", "页面正常加载", "1. 访问Google首页\n2. 定位搜索输入框\n3. 输入搜索关键词", "搜索输入框存在，可正常输入文字"],
        ["TC-005", "验证搜索按钮存在", "搜索", "P0", "页面正常加载", "1. 访问Google首页\n2. 定位Google搜索按钮", "Google搜索按钮存在且可见"],
        ["TC-006", "验证I'm Feeling Lucky按钮存在", "搜索", "P1", "页面正常加载", "1. 访问Google首页\n2. 定位I'm Feeling Lucky按钮", "I'm Feeling Lucky按钮存在且可见"],
        ["TC-007", "验证输入关键词后执行搜索", "搜索", "P0", "网络正常", "1. 访问Google首页\n2. 输入搜索关键词\n3. 点击搜索按钮", "页面跳转到搜索结果页"],
        ["TC-008", "验证按Enter键执行搜索", "搜索", "P0", "网络正常", "1. 访问Google首页\n2. 输入搜索关键词\n3. 按Enter键", "页面跳转到搜索结果页"],
        ["TC-009", "验证空搜索提交", "搜索", "P1", "无", "1. 访问Google首页\n2. 不输入任何关键词\n3. 点击搜索按钮", "页面不跳转或显示提示"],
        ["TC-010", "验证语音搜索按钮存在", "搜索", "P1", "页面正常加载", "1. 访问Google首页\n2. 查找语音搜索按钮", "语音搜索按钮存在"],
        ["TC-011", "验证图片搜索按钮存在", "搜索", "P1", "页面正常加载", "1. 访问Google首页\n2. 查找图片搜索按钮", "图片搜索按钮存在"],
        ["TC-012", "验证顶部导航栏显示", "导航", "P0", "页面正常加载", "1. 访问Google首页\n2. 查看顶部导航栏", "导航栏正常显示"],
        ["TC-013", "验证Gmail链接存在", "导航", "P0", "页面正常加载", "1. 访问Google首页\n2. 查找Gmail链接", "Gmail链接存在且可点击"],
        ["TC-014", "验证图片链接存在", "导航", "P0", "页面正常加载", "1. 访问Google首页\n2. 查找图片链接", "图片链接存在且可点击"],
        ["TC-015", "验证应用网格按钮存在", "导航", "P1", "页面正常加载", "1. 访问Google首页\n2. 查找应用网格按钮", "应用网格按钮存在"],
        ["TC-016", "验证用户头像按钮存在", "导航", "P1", "页面正常加载", "1. 访问Google首页\n2. 查找用户头像按钮", "用户头像按钮存在"],
        ["TC-017", "验证底部设置链接", "页脚", "P1", "页面正常加载", "1. 访问Google首页\n2. 滚动到页面底部\n3. 查找设置链接", "设置链接存在"],
        ["TC-018", "验证底部隐私权政策链接", "页脚", "P1", "页面正常加载", "1. 访问Google首页\n2. 滚动到页面底部\n3. 查找隐私权政策链接", "隐私权政策链接存在"],
        ["TC-019", "验证底部服务条款链接", "页脚", "P1", "页面正常加载", "1. 访问Google首页\n2. 滚动到页面底部\n3. 查找服务条款链接", "服务条款链接存在"],
        ["TC-020", "验证底部广告链接", "页脚", "P2", "页面正常加载", "1. 访问Google首页\n2. 滚动到页面底部\n3. 查找广告链接", "广告相关链接存在"],
        ["TC-021", "验证搜索建议功能", "搜索", "P0", "网络正常", "1. 访问Google首页\n2. 输入搜索关键词\n3. 等待搜索建议出现", "搜索建议正常显示"],
        ["TC-022", "验证搜索历史记录", "搜索", "P1", "有搜索历史", "1. 访问Google首页\n2. 点击搜索输入框\n3. 查看搜索历史", "搜索历史记录正常显示"],
        ["TC-023", "验证I'm Feeling Lucky随机跳转", "搜索", "P1", "网络正常", "1. 访问Google首页\n2. 点击I'm Feeling Lucky按钮", "跳转到随机Google服务页面"],
        ["TC-024", "验证多语言切换", "功能", "P2", "页面正常加载", "1. 访问Google首页\n2. 查找语言设置\n3. 切换语言", "页面语言切换成功"],
        ["TC-025", "验证页面响应式设计", "界面", "P2", "使用不同尺寸浏览器", "1. 使用不同尺寸浏览器窗口访问\n2. 查看页面布局", "页面布局自适应屏幕尺寸"],
    ]

    for row_idx, case in enumerate(test_cases, 2):
        for col_idx, value in enumerate(case, 1):
            ws_cases.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, len(headers) + 1):
        ws_cases.column_dimensions[get_column_letter(col)].width = 20

    ws_config = wb.create_sheet("配置")

    config_data = [
        ["配置项", "值"],
        ["URL", "https://www.google.com/"],
        ["登录账号", "无"],
        ["登录密码", "无"],
        ["测试描述", "主页的所有功能"],
    ]

    for row_idx, row_data in enumerate(config_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws_config.column_dimensions['A'].width = 15
    ws_config.column_dimensions['B'].width = 40

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_dir = os.path.join(base_dir, "test_cases", "google")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "google.xlsx")
    wb.save(output_path)
    print(f"测试用例已生成到: {output_path}")


if __name__ == "__main__":
    create_google_home_test_cases()
