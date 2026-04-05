import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def create_google_home_test_cases():
    wb = openpyxl.Workbook()

    ws_cases = wb.active
    ws_cases.title = "测试用例"

    headers = ["用例编号", "用例名称", "模块/功能", "优先级", "预置条件", "测试步骤", "预期结果"]
    for col, header in enumerate(headers, 1):
        cell = ws_cases.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    test_cases = [
        ["TC-001", "Google Logo显示", "页面元素", "P0", "网络连接正常", "1. 打开浏览器\n2. 访问 https://www.google.com/\n3. 等待页面加载完成", "Google Logo正常显示"],
        ["TC-002", "搜索输入框存在", "搜索功能", "P0", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查找搜索输入框", "搜索输入框可见且可编辑"],
        ["TC-003", "搜索按钮存在", "搜索功能", "P0", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查找搜索按钮", "搜索按钮可见且可点击"],
        ["TC-004", "Google搜索功能验证", "搜索功能", "P0", "网络连接正常", "1. 访问 https://www.google.com/\n2. 在搜索框输入\"test\"\n3. 点击搜索按钮或按回车", "搜索结果页面正常加载，显示相关结果"],
        ["TC-005", "I'm Feeling Lucky按钮", "搜索功能", "P1", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查找I'm Feeling Lucky按钮", "按钮可见且可点击"],
        ["TC-006", "顶部导航栏链接", "页面元素", "P1", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查找顶部导航栏链接（Gmail、图片等）", "导航链接正常显示"],
        ["TC-007", "页面标题验证", "页面功能", "P0", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 获取页面标题", "页面标题为Google或Google Home"],
        ["TC-008", "搜索框占位符文本", "搜索功能", "P1", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 检查搜索框占位符", "显示\"Google 搜索或输入网址\"或类似文本"],
        ["TC-009", "多语言切换", "页面功能", "P2", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查找语言/地区设置", "可选择不同语言和地区"],
        ["TC-010", "页面响应时间", "性能测试", "P1", "网络连接正常", "1. 记录开始时间\n2. 访问 https://www.google.com/\n3. 记录页面加载完成时间", "页面加载时间在3秒以内"],
        ["TC-011", "底部链接显示", "页面元素", "P2", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 滚动至页面底部", "底部显示相关链接（关于、广告、业务等）"],
        ["TC-012", "搜索建议功能", "搜索功能", "P1", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 在搜索框输入部分关键词\n3. 等待自动补全建议", "显示搜索建议下拉菜单"],
        ["TC-013", "页面在不同浏览器显示", "兼容性", "P2", "网络连接正常", "1. 使用Chrome、Firefox、Edge浏览器访问\n2. 检查页面布局", "各浏览器页面显示正常"],
        ["TC-014", "搜索框回车键提交", "搜索功能", "P0", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 在搜索框输入关键词\n3. 按Enter键", "执行搜索并跳转到结果页面"],
        ["TC-015", "页面版权信息", "页面元素", "P2", "已打开Google首页", "1. 访问 https://www.google.com/\n2. 查看页面底部", "显示© 2026 Google LLC版权信息"],
    ]

    for row_idx, case in enumerate(test_cases, 2):
        for col_idx, value in enumerate(case, 1):
            cell = ws_cases.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, len(headers) + 1):
        ws_cases.column_dimensions[get_column_letter(col)].width = 20

    ws_config = wb.create_sheet("配置")

    config_data = [
        ["配置项", "值"],
        ["URL", "https://www.google.com/"],
        ["登录账号", "无"],
        ["登录密码", "无"],
        ["测试描述", "主页的所有功能"],
    ]

    for row_idx, row_data in enumerate(config_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws_config.column_dimensions['A'].width = 15
    ws_config.column_dimensions['B'].width = 40

    import os
    os.makedirs("test_cases", exist_ok=True)

    output_path = "test_cases/google.xlsx"
    wb.save(output_path)
    print(f"测试用例已生成到: {output_path}")


if __name__ == "__main__":
    create_google_home_test_cases()
